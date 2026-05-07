from __future__ import annotations

"""
Import vendor list-price Excel files into the database.

Expected Excel layout (Salesforce format):
  Row 1-2: FX rate metadata
  Row 3:   Headers — Vendor, Description, Manufacturer Part Number,
           List Price in US$, NASPO Price, GBP, AUD, CAD, AED
  Row 4+:  Data rows
"""

import re
from datetime import datetime, date
from sqlalchemy.orm import Session
import openpyxl

from models import PriceList, ListPrice
from vendor_normalization import normalize_line_item


_DURATION_RE = re.compile(r'(\d+)\s*[-]?\s*[Yy]ear', re.IGNORECASE)


def _parse_duration_years(description: str) -> float:
    m = _DURATION_RE.search(description)
    if m:
        return float(m.group(1))
    return 1.0


def _parse_effective_date(filename: str) -> date | None:
    m = re.search(r'(\d{1,2})\w{0,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+(\d{4})', filename, re.IGNORECASE)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)[:3]} {m.group(3)}", "%d %b %Y").date()
        except ValueError:
            pass
    return None


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def import_list_prices(
    file_path: str,
    filename: str,
    db: Session,
    uploaded_by: int | None = None,
) -> dict:
    """
    Parse an Excel price-book and store rows in list_prices.

    Returns summary dict with import stats.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Parse FX metadata from rows 1-2
    fx_metadata = {}
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    fx_labels = ["gbp", "aud", "cad", "aed"]
    for i, label in enumerate(fx_labels):
        val = row2[5 + i] if len(row2) > 5 + i else None
        if val is not None:
            fx_metadata[label] = float(val)

    # Detect vendor from first data row
    first_data = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    raw_vendor = (first_data[0] or "").strip()
    canonical_vendor, _ = normalize_line_item(raw_vendor, "", db)

    # Deactivate previous price lists for this vendor
    db.query(PriceList).filter(
        PriceList.vendor_name == canonical_vendor,
        PriceList.is_active == True,
    ).update({"is_active": False})

    effective = _parse_effective_date(filename)

    price_list = PriceList(
        vendor_name=canonical_vendor,
        source_filename=filename,
        effective_date=effective,
        fx_rate_metadata=fx_metadata,
        is_active=True,
        uploaded_by=uploaded_by,
    )
    db.add(price_list)
    db.flush()

    imported = 0
    skipped = 0
    unmatched_products = set()

    for row in ws.iter_rows(min_row=4, values_only=True):
        vendor_raw = (row[0] or "") if row[0] else ""
        description = (row[1] or "") if row[1] else ""
        part_number = row[2] if len(row) > 2 else None
        list_price = _safe_float(row[3] if len(row) > 3 else None)
        naspo = _safe_float(row[4] if len(row) > 4 else None)
        gbp = _safe_float(row[5] if len(row) > 5 else None)
        aud = _safe_float(row[6] if len(row) > 6 else None)
        cad = _safe_float(row[7] if len(row) > 7 else None)
        aed = _safe_float(row[8] if len(row) > 8 else None)

        if not description.strip():
            skipped += 1
            continue

        # Normalize part number to string
        if part_number is not None:
            part_number = str(part_number).strip()
            if part_number.endswith(".0"):
                part_number = part_number[:-2]

        duration = _parse_duration_years(description)
        annual_price = list_price / duration if duration > 0 else list_price

        _, canonical_product = normalize_line_item(canonical_vendor, description, db)

        # Track unmatched (fallback title-cased names)
        if canonical_product == description.strip().title() or canonical_product == description.strip():
            unmatched_products.add(description.strip()[:80])

        lp = ListPrice(
            price_list_id=price_list.id,
            vendor_name=canonical_vendor,
            product_name_raw=description.strip(),
            product_name_canonical=canonical_product,
            manufacturer_part_number=part_number,
            list_price_usd=list_price,
            naspo_price=naspo if naspo else None,
            list_price_gbp=gbp if gbp else None,
            list_price_aud=aud if aud else None,
            list_price_cad=cad if cad else None,
            list_price_aed=aed if aed else None,
            duration_years=duration,
            list_price_annual_usd=round(annual_price, 2),
        )
        db.add(lp)
        imported += 1

    price_list.total_rows_imported = imported
    db.commit()
    wb.close()

    return {
        "price_list_id": price_list.id,
        "vendor_name": canonical_vendor,
        "effective_date": str(effective) if effective else None,
        "rows_imported": imported,
        "rows_skipped": skipped,
        "unmatched_canonical_count": len(unmatched_products),
        "sample_unmatched": sorted(unmatched_products)[:20],
    }
